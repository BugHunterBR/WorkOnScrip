import json
import os
from openpyxl import load_workbook

# Carregar configurações do JSON
with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'assistant', 'configPDFC.json'), 'r') as config_file:
    config = json.load(config_file)

email_field = config.get('email_field')
path_file = config.get('path_file')
sheet = config.get('sheet')
MarckCheck = int(config.get('MarckCheck', 1))
MarckFlag = int(config.get('MarckFlag', 2))

file = load_workbook(path_file)

select_sheet = file[sheet]
print(select_sheet)

#for line in range(1 ,select_sheet.max_row + 1):
#    print(line)

print(len(select_sheet['A']))