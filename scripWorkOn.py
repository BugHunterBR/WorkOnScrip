import os
import json
import logging as log
import win32com.client as client
import re
from tqdm import tqdm
from openpyxl import load_workbook

# --proxy="http://rb-proxy-de.bosch.com:8080"

log.basicConfig(
    filename=os.path.join(os.path.dirname(os.path.abspath(__file__)), 'assistant', 'Processamento_PDFC.log'),
    level=log.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Carregar configurações do JSON
with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'assistant', 'configPDFC.json'), 'r') as config_file:
    config = json.load(config_file)

email_field = config.get('email_field')
path_file = config.get('path_file')
sheet = config.get('sheet')
MarckCheck = int(config.get('MarckCheck', 1))
MarckFlag = int(config.get('MarckFlag', 2))

try:
    file = load_workbook(path_file)
    select_sheet = file[sheet]

    outlook = client.Dispatch('Outlook.Application')
    namespace = outlook.GetNamespace('MAPI')
    inbox = namespace.Folders[email_field].Folders['WorkOn']

    messages = inbox.Items
    messages.Sort("[ReceivedTime]", True)

    try:
        for msg in tqdm(messages, desc='Processando e-mails', unit=' e-mail'):
            sender = msg.SenderEmailAddress

            if (
                msg.Class == 43 
                and msg.FlagStatus == 0 
                and sender == 'NoReply.Workon@bosch.com' 
                and re.match(r'Ação Requerida \[Substituted\]\s+(.*?)\s+- CR_\d+_Cadastro Novo Item', msg.Subject)
            ):
                try:
                    subject = msg.Subject
                    body = msg.Body

                    standard_subject = r'Ação Requerida \[Substituted\]\s+(.*?)\s+- CR_\d+_Cadastro Novo Item'
                    match_subject = re.search(standard_subject, subject)

                    standard_body = r'Descrição\s+([\s\S]*?)\s+Iniciado por'
                    match_body = re.search(standard_body, body)

                    if match_subject and match_body:
                        result_subject = match_subject.group(1).strip()
                        print(f"\nTexto capturado do assunto: {result_subject}")

                        result_body = match_body.group(1).strip()

                        standard_result = r":\s*([^|]+)\s*(?:\|\||$)"
                        values = [value.strip() for value in re.findall(standard_result, result_body)]

                        next_row = select_sheet.max_row + 1

                        select_sheet.cell(row=next_row, column=1, value=result_subject)

                        for col, value in enumerate(values, start=2):
                            select_sheet.cell(row=next_row, column=col, value=value)
                        
                        '''
                        msg.MarkAsTask(MarckCheck)
                        msg.FlagStatus = MarckCheck
                        msg.save()
                        
                        msg.Move(namespace.Folders[email_field].Folders['WorkOn_Processed'])
                        '''

                    '''
                    if msg.FlagStatus != MarckCheck:
                        msg.MarkAsTask(MarckFlag)
                        msg.FlagStatus = MarckFlag
                        msg.save()
                        
                        msg.Move(namespace.Folders[email_field].Folders['WorkOn_Alert'])
                    '''

                except Exception as e:
                    log.error(f'Erro ao processar padrão de corpo e/ou texto do e-mail: {e}', exc_info=True)
            
        file.save(path_file)
        
    except Exception as e:
        log.error(f'Erro no processamento dos e-mails: {e}', exc_info=True)

except Exception as e:
    log.error(f'Erro geral: {e}', exc_info=True)